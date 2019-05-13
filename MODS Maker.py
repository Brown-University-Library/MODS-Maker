import os, csv, xlsxwriter
from lxml import etree
from openpyxl import load_workbook
import openpyxl
from lxml.builder import ElementMaker
import string, codecs
import chardet
import datetime
import re
import xlrd
from copy import copy

requiredcolumns = ["subjectTopicsFAST","Ignore", "fileTitle", "itemTitle", "subTitle", "place","dateText","dateStart","dateEnd","dateBulkStart","dateBulkEnd","dateQualifier", "shelfLocator1", "shelfLocator1ID", "shelfLocator2", "shelfLocator2ID", "shelfLocator3","shelfLocator3ID","typeOfResource","genreAAT","genreLCSH","genreLocal","genreRBGENR","extentQuantity","extentSize","extentSpeed","form","noteScope","noteHistorical","noteHistoricalClassYear","noteGeneral","language","noteAccession","identifierBDR","publisher","namePersonCreatorLC","namePersonCreatorLocal","nameCorpCreatorLC","nameCorpCreatorLocal","namePersonOtherLC","namePersonOtherLocal","subjectNamesLC","subjectNamesLocal","subjectCorpLC","subjectCorpLocal","subjectTopicsLC","subjectTopicsLocal","subjectGeoLC","subjectTemporalLC","subjectTitleLC","collection","dateTextParent","callNumber","repository","findingAid","digitalOrigin","rightsStatementText","rightsStatementURI", "useAndReproduction", "coordinates", "scale", "projection"]
langcode = {}
langcodeopp = {}
scriptcode = {}
langissue = False

def getSplitCharacter(string):
    if ";" in string:
        return(";")
    else:
        return("|")

def messageToUser(messagetitle, message):
    print("")
    print(messagetitle)
    print(message)
    try:
        raw_input("Press Enter to continue . . .")
    except SyntaxError:
        print("Syntax Error")
    except TypeError:
        print("Type Error")

def multilinefield(parentelement, originalfieldname, eadfieldname):
    newelement = etree.SubElement(parentelement, eadfieldname)
    lines = cldata.get(originalfieldname, '').splitlines()
    for line in lines:
        pelement = etree.SubElement(newelement, "p")
        pelement.text = ' '.join(line.split()).decode('utf-8')

def repeatingfield(parentelement, refdict, originalfieldname, modsfieldname, modsattributes, subject, subjectattributes):
    splitcharacter = ""
    originalparentelement = parentelement

    if ";" in refdict.get(originalfieldname, ''):
        splitcharacter = ";"
    else:
        splitcharacter = "|"

    for namesindex, addedentry in enumerate(refdict.get(originalfieldname, '').split(splitcharacter)):

        if subject == True:
            subjectelement = etree.SubElement(parentelement, "{http://www.loc.gov/mods/v3}subject", subjectattributes)
            parentelement = subjectelement

        namecontrolaccesselement = etree.SubElement(parentelement, modsfieldname, modsattributes)
        namecontrolaccesselement.text = ' '.join(addedentry.replace("|d", "").replace("|e", "").split()).decode('utf-8')

        parentelement = originalparentelement

def repeatingnamefield(parentelement, refdict, originalfieldname, topmodsattributes, predefinedrole, subject, splitcharacter):
    originalparentelement = parentelement
    if splitcharacter == 'v':
        if ";" in refdict.get(originalfieldname, ''):
            splitcharacter = ";"
        else:
            splitcharacter = "|"

    for nameindex, name in enumerate(refdict.get(originalfieldname, '').split(splitcharacter)):
        nametext = ""
        datetext = ""
        roletext = predefinedrole


        for textindex, text in enumerate(name.split(',')):
            textrevised = ' '.join(text.split()).replace('|d', '').replace('|e','')

            if textrevised == '':
                continue

            max_index = len(xmltext(name).split(','))-1

            if textindex == 0:
                nametext = nametext + textrevised + ", "
            elif hasYear(textrevised) == True:
                datetext = datetext + textrevised
            elif isAllLower(textrevised) == True:
                roletext = text
            elif hasLetters(textrevised) != None:
                nametext = nametext + textrevised + " "

        if nametext == '':
            continue

        if subject == True:
            subjectelement = etree.SubElement(parentelement, "{http://www.loc.gov/mods/v3}subject")
            parentelement = subjectelement

        nameelement = etree.SubElement(parentelement, "{http://www.loc.gov/mods/v3}name", topmodsattributes)
        namepart = etree.SubElement(nameelement, "{http://www.loc.gov/mods/v3}namePart")
        namepart.text = xmltext(nametext).rstrip(',')
        namedatepart = etree.SubElement(nameelement, "{http://www.loc.gov/mods/v3}namePart", {"type":"date"})
        namedatepart.text = xmltext(datetext).lstrip(',').rstrip(',').replace('|d','')
        modsrole = etree.SubElement(nameelement, "{http://www.loc.gov/mods/v3}role")
        modsroleterm = etree.SubElement(modsrole, "{http://www.loc.gov/mods/v3}roleTerm", {"type":"text", "authority":"marcrelator"})
        modsroleterm.text = xmltext(roletext).lstrip(',').rstrip(',').replace('|e','')

        parentelement = originalparentelement

def xmltext(text):
    text = text.replace('\n', ' ').replace('\r', ' ')
    text = text.replace('<title>', '').replace('</title>', '')
    text = text.replace('<geogname>', '- ').replace('</geogname>', '')
    return(' '.join(str(text).split()).decode('utf-8'))

def convertEncoding(from_encode,to_encode,old_filepath,target_file):
    f1=open(old_filepath)
    content2=[]
    while True:
        line=f1.readline()
        content2.append(line.decode(from_encode).encode(to_encode))
        if len(line) ==0:
            break

    f1.close()
    f2=open(target_file,'w')
    f2.writelines(content2)
    f2.close()

def hasNumbers(s):
    return any(i.isdigit() for i in s)

def hasYear(s):
    numbercount = 0
    for i in s:
        if i.isdigit() == True:
            numbercount = numbercount + 1
    if numbercount > 3:
        return True
    else:
        return False

def hasLetters(s):
    return re.search('[a-zA-Z]', s)

def isAllLower(s):
    nonlowercase = 0
    for i in s.replace(' ', ''):
        if i.islower() == False:
            nonlowercase = nonlowercase + 1
            break
    if nonlowercase > 0:
        return False
    else:
        return True



def let_user_pick(message, options):
    print("")
    print(message)
    for idx, element in enumerate(options):
        print("{}) {}".format(idx+1,element))
    i = input("Enter number: ")
    try:
        if 0 < int(i) <= len(options):
            return options[i-1]
    except:
        pass
    return None

def XLSDictReader(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        rowarray = []

        for row in range(1, sheet.nrows):
            rowdictionary = {}
            for column in range(sheet.ncols):
                #If the value is a number, turn it into a string.
                newvalue = ''
                if sheet.cell(row,column).ctype > 1:
                    newvalue = str(sheet.cell_value(row,column)).replace('|d', '').replace('|e', '').replace('|',';').encode('utf-8')
                else:
                    newvalue = sheet.cell_value(row,column).replace('|d', '').replace('|e', '').replace('|',';').encode('utf-8')

                #If the column is repeating, serialize the row values.
                if rowdictionary.get(sheet.cell_value(0,column), '') != '':
                    rowdictionary[sheet.cell_value(0,column)] = rowdictionary[sheet.cell_value(0,column)] + getSplitCharacter(rowdictionary[sheet.cell_value(0,column)]) + newvalue
                else:
                    rowdictionary[sheet.cell_value(0,column)] = newvalue
            rowarray.append(rowdictionary)
        return(rowarray)

def XLSDictReaderLanguageCode(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        langcode = {}

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 0).encode('utf-8')
            value = sheet.cell_value(row, 1).encode('utf-8')
            langcode[key] = value
        return(langcode)

def XLSDictReaderLanguageCodeOpp(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        langcode = {}

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 1).encode('utf-8')
            value = sheet.cell_value(row, 0).encode('utf-8')
            langcode[key] = value
        return(langcode)

def XLSDictReaderScriptCode(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        scriptcode = {}

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 0).encode('utf-8')
            value = sheet.cell_value(row, 2)
            scriptcode[key] = value
        return(scriptcode)

print("._. MODS Maker ._.")

#Get all languages codes and script codes.
langcode = XLSDictReaderLanguageCode(os.getcwd() + "/data/SupportedLanguages.xlsx","languages xlsx")
langcodeopp = XLSDictReaderLanguageCodeOpp(os.getcwd() + "/data/SupportedLanguages.xlsx","languages xlsx")
scriptcode = XLSDictReaderScriptCode(os.getcwd() + "/data/SupportedLanguages.xlsx","languages xlsx")

#Make the cache directory if it doesn't yet exist.
try:
     os.mkdir(os.getcwd() + '/'+ 'cache')
except OSError:
     print ("" )
else:
     print ("")

#Get a list of CSV/XLSX/XLS files in the directory and present them to the user.
csvfilelist = []
chosenfile = ''

for file in os.listdir(os.getcwd()):
    if file.endswith(".csv") or file.endswith(".xlsx") or file.endswith(".xls"):
        csvfilelist.append(file)

if len(csvfilelist) > 1:
    chosenfile = let_user_pick("Please choose a file:", csvfilelist)
    print(chosenfile + " was selected.")
else:
    chosenfile = csvfilelist[0]

#Extract spreadsheet data to csvdata dictionary.
csvdata = {}

if chosenfile.endswith('.csv') == True:
    #Convert the file to UTF-8.
    convertFile = open(chosenfile,'r')
    data = convertFile.read()
    convertFile.close()

    print(chardet.detect(data))

    convertEncoding(chardet.detect(data)['encoding'], "utf-8", chosenfile, os.getcwd() + "/cache/cacheut8.csv")
    csvdata = csv.DictReader(open(os.getcwd() + "/cache/cacheut8.csv"))
elif chosenfile.endswith('.xlsx') == True or chosenfile.endswith('.xls') == True:
    excel = xlrd.open_workbook(chosenfile)
    sheetnames = excel.sheet_names()
    chosensheet = let_user_pick("Please choose the Excel sheet you'd like to process:", sheetnames)

    #Get column headers and check them against required columns.
    selectedsheet = excel.sheet_by_name(chosensheet)
    columnsinsheet = [str(cell.value) for cell in selectedsheet.row(0)]

    print("")
    missingcolumns = []
    for column in requiredcolumns:
        if (column in columnsinsheet) == False:
            print("Missing spreadsheet column: " + column)
            missingcolumns.append(column)

    if len(missingcolumns) != 0:
        messageToUser("*Missing Columns Detected*", "The columns above are missing from your spreadsheet. The script will now continue without them.")
        originalfile = chosenfile

    csvdata = XLSDictReader(chosenfile, chosensheet)
    chosenfile = chosensheet

#Create the output directory and save the path to the output_path variable.
now = datetime.datetime.now()
output_path = os.getcwd()

try:
     os.mkdir(os.getcwd() + '/'+ chosenfile + " " + now.strftime("%m-%d-%Y %H %M " + str(now.second)))
except OSError:
     print ("" )
else:
     print ("")
     output_path = os.getcwd() + '/'+ chosenfile + " " + now.strftime("%m-%d-%Y %H %M " + str(now.second))

#Create the error CSV.
errorfile = open(output_path + '/Error Report ' + now.strftime("%m-%d-%Y %H %M " + str(now.second)) + '.csv', mode='wb')
errorcsvwriter = csv.writer(errorfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
errorcsvwriter.writerow(['Spreadsheet Row', 'BDR Number', 'Column Name', 'Column Contents', 'Potential Issue'])

#Set up namespaces and attributes for XML.
attr_qname = etree.QName("http://www.w3.org/2001/XMLSchema-instance", "schemaLocation")
ns_map = {"mods" : "http://www.loc.gov/mods/v3", "xsi" : "http://www.w3.org/2001/XMLSchema-instance", "xlink" : "http://www.w3.org/1999/xlink"}

#root = etree.Element("{http://www.loc.gov/mods/v3}modsCollection", {attr_qname: "http://www.loc.gov/mods/v3 http://www.loc.gov/mods/v3/mods-3-7.xsd"}, nsmap=ns_map)
#print(etree.tostring(root))

amountofrecords = 0
rowindex = 2

#Create a MODS file for every row in the input CSV file.
for row in csvdata:

    #Ignore rows that contain EAD-specific data or anything in the Ignore column.
    if row.get('seriesTitle', '') != '':
        continue
    if row.get('subSeriesTitle', '') != '':
        continue
    if row.get('Ignore', '') != '':
        continue

    #Set up the top-level mods element.
    modstop = etree.Element("{http://www.loc.gov/mods/v3}mods", {attr_qname: "http://www.loc.gov/mods/v3 http://www.loc.gov/mods/v3/mods-3-7.xsd"}, nsmap=ns_map)

    #mods:titleInfo
    titleinfo = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}titleInfo")
    title = etree.SubElement(titleinfo, "{http://www.loc.gov/mods/v3}title")
    if row.get("fileTitle", '') != "":
        title.text = xmltext(row.get("fileTitle", ''))
    #elif row.get("title", '') != "":
    #    title.text = xmltext(row.get("title", ''))
    else:
        title.text = xmltext(row.get("itemTitle", ''))
    #title.text = xmltext(row.get("title"]) #' '.join(row["title", '').split()).decode('utf-8')
    subtitle = etree.SubElement(titleinfo, "{http://www.loc.gov/mods/v3}subTitle")
    subtitle.text = xmltext(row.get("subTitle", ''))
    # ' '.join(row.get("subTitle", '').split())

    #namePersonCreatorLC
    repeatingnamefield(modstop, row, 'namePersonCreatorLC', {"type":"personal", "authority":"naf"}, 'creator', False, 'v')
    repeatingnamefield(modstop, row, 'namePersonCreatorLocal', {"type":"personal", "authority":"local"}, 'creator', False, 'v')
    repeatingnamefield(modstop, row, 'namePersonOtherLC', {"type":"personal", "authority":"naf"}, '', False, 'v')
    repeatingnamefield(modstop, row, 'namePersonOtherLocal', {"type":"personal", "authority":"local"}, '', False, 'v')
    repeatingnamefield(modstop, row, 'nameCorpCreatorLC', {"type":"corporate", "authority":"naf"}, 'creator', False, 'v')
    repeatingnamefield(modstop, row, 'nameCorpCreatorLocal', {"type":"corporate", "authority":"local"}, 'creator', False, 'v')


    #typeOfResource
    typeofresource = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}typeOfResource")
    typeofresource.text = ' '.join(row.get("typeOfResource", '').split())

    #genre
    repeatingfield(modstop, row, "genreAAT", "{http://www.loc.gov/mods/v3}genre", {"authority":"aat"}, False, {})
    repeatingfield(modstop, row, "genreLCSH", "{http://www.loc.gov/mods/v3}genre", {"authority":"lcsh"}, False, {})
    repeatingfield(modstop, row, "genreLocal", "{http://www.loc.gov/mods/v3}genre", {"authority":"local"}, False, {})
    repeatingfield(modstop, row, "genreRBGENR", "{http://www.loc.gov/mods/v3}genre", {"authority":"rbgenr"}, False, {})

    #note
    notescopeelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}abstract", {"type":"general", "displayLabel":"Scope and Contents note"})
    notescopeelement.text = xmltext(row.get("noteScope", ''))
    #' '.join(row.get("noteScope", '').split()).decode('utf-8')

    noteGeneralelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}note", {"type":"general"})
    noteGeneralelement.text = xmltext(row.get("noteGeneral", ''))
    #' '.join(row.get("noteGeneral", '').split()).decode('utf-8')

    noteAccessionelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}note", {"type":"acquisition", "displayLabel":"Immediate form of acquisition"})
    noteAccessionelement.text = xmltext(row.get("noteAccession", ''))
    #' '.join(row.get("noteAccession", '').split()).decode('utf-8')

    noteHistoricalelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}note", {"type":"biographical/historical", "displayLabel":"Biographical note"})
    noteHistoricalelement.text = xmltext(row.get("noteHistorical", ''))
    #' '.join(row.get("noteHistorical", '').split()).decode('utf-8')

    noteHistoricalClassYearelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}note", {"type":"biographical/historical", "displayLabel":"Class year"})
    noteHistoricalClassYearelement.text = xmltext(row.get("noteHistoricalClassYear", ''))
    # ' '.join(row.get("noteHistoricalClassYear", '').split()).decode('utf-8')

    notePreferredCitation = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}note", {"type":"preferredCitation"})
    notePreferredCitationstring = title.text # ' '.join(row.get("title", '').split()).rstrip('.').decode('utf-8')
    if row.get("collection", '') != "":
        notePreferredCitationstring = notePreferredCitationstring + ", " + ' '.join(row.get("collection", '').split()).decode('utf-8')
    if row.get("callNumber", '') != "":
        notePreferredCitationstring = notePreferredCitationstring + ", " + ' '.join(row.get("callNumber", '').split())
    notePreferredCitation.text = notePreferredCitationstring + ', Brown University Library'

    #originInfo
    originInfoelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}originInfo")

    publisherelement = etree.SubElement(originInfoelement, "{http://www.loc.gov/mods/v3}publisher")
    publisherelement.text = ' '.join(row.get("publisher", '').split())

    dateQualifierAttribute = {}

    if row.get("dateQualifier", '') != "":
        dateQualifierAttribute = {"qualifier": row.get("dateQualifier", '')}

    dateCreatedelement = etree.SubElement(originInfoelement, "{http://www.loc.gov/mods/v3}dateCreated", dateQualifierAttribute)
    dateCreatedelement.text = ' '.join(row.get("dateText", '').split()).replace('.0','')

    dateStartelementdict = {"encoding":"w3cdtf", "keyDate":"yes", "point":"start"}
    dateStartelementdict.update(dateQualifierAttribute)
    dateStartelement = etree.SubElement(originInfoelement, "{http://www.loc.gov/mods/v3}dateCreated", dateStartelementdict)
    dateStartelement.text = ' '.join(str(row.get("dateStart", '')).split()).replace('.0','')

    dateEndelementdict = {"encoding":"w3cdtf", "point":"end"}
    dateEndelementdict.update(dateQualifierAttribute)
    dateEndelement = etree.SubElement(originInfoelement, "{http://www.loc.gov/mods/v3}dateCreated", dateEndelementdict)
    dateEndelement.text = ' '.join(str(row.get("dateEnd", '')).split()).replace('.0','')

    placeelement = etree.SubElement(originInfoelement, "{http://www.loc.gov/mods/v3}place")
    placeTermelement = etree.SubElement(placeelement, "{http://www.loc.gov/mods/v3}placeTerm", {"type":"text"})
    placeTermelement.text = ' '.join(row.get("place", '').split())

    #language
    languagesplitcharacter = getSplitCharacter(row.get("language", ''))
    for language in row.get("language", '').split(languagesplitcharacter):
        languageelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}language")
        languageTermelement = etree.SubElement(languageelement, "{http://www.loc.gov/mods/v3}languageTerm", {"type":"code", "authority":"iso639-2b"})

        if len(xmltext(language)) > 3:
            if xmltext(language) in langcode:
                 languageTermelement.text = langcode[xmltext(language)]
            else:
                 languageTermelement.text = ' '.join(language.split())
                 langissue = True
        else:
            languageTermelement.text = xmltext(language)

    #physicalDescription
    physicalDescriptionelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}physicalDescription")

    extentQuantityelement = etree.SubElement(physicalDescriptionelement, "{http://www.loc.gov/mods/v3}extent")
    extentQuantityelement.text = ' '.join(row.get("extentQuantity", '').split())

    extentSizeelement = etree.SubElement(physicalDescriptionelement, "{http://www.loc.gov/mods/v3}extent")
    extentSizeelement.text = ' '.join(row.get("extentSize", '').split())

    extentSpeedelement = etree.SubElement(physicalDescriptionelement, "{http://www.loc.gov/mods/v3}extent")
    extentSpeedelement.text = ' '.join(row.get("extentSpeed", '').split())

    digitalOriginelement = etree.SubElement(physicalDescriptionelement, "{http://www.loc.gov/mods/v3}digitalOrigin")
    digitalOriginelement.text = ' '.join(row.get("digitalOrigin", '').split())

    formelement = etree.SubElement(physicalDescriptionelement, "{http://www.loc.gov/mods/v3}form")
    formelement.text = ' '.join(row.get("form", '').split())

    #accessCondition
    useAndReproductionelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}accessCondition", {"type":"useAndReproduction"})
    useAndReproductionelement.text = ' '.join(row.get("useAndReproduction", '').split())

    rightsStatementelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}accessCondition", {"type":"rightsStatement","{http://www.w3.org/1999/xlink}href":xmltext(row.get("rightsStatementURI", ''))})
    rightsStatementelement.text = ' '.join(row.get("rightsStatementText", '').split())

    restrictionOnAccesselement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}accessCondition", {"type":"restrictionOnAccess"})
    restrictionOnAccesselement.text = "Collection is open for research."

    #subject
    repeatingnamefield(modstop, row, 'subjectNamesLC', {"type":"personal", "authority":"naf"}, '', True, 'v')
    repeatingnamefield(modstop, row, 'subjectNamesLocal', {"type":"personal", "authority":"local"}, '', True, 'v')
    repeatingnamefield(modstop, row, 'subjectCorpLC', {"type":"corporate", "authority":"naf"}, '', True, 'v')
    repeatingnamefield(modstop, row, 'subjectCorpLocal', {"type":"corporate", "authority":"local"}, '', True, 'v')

    repeatingfield(modstop, row, "subjectTopicsLC", "{http://www.loc.gov/mods/v3}topic", {}, True, {"authority":"lcsh"})
    repeatingfield(modstop, row, "subjectTopicsLocal", "{http://www.loc.gov/mods/v3}topic", {}, True, {"authority":"local"})
    repeatingfield(modstop, row, "subjectTopicsFAST", "{http://www.loc.gov/mods/v3}topic", {}, True, {"authority":"fast"})
    repeatingfield(modstop, row, "subjectGeoLC", "{http://www.loc.gov/mods/v3}geographic", {}, True, {"authority":"lcsh"})
    repeatingfield(modstop, row, "subjectTemporalLC", "{http://www.loc.gov/mods/v3}temporal", {}, True, {"authority":"lcsh"})
    repeatingfield(modstop, row, "subjectTitleLC", "{http://www.loc.gov/mods/v3}title", {}, True, {"authority":"naf"})

    #cartographic
    subjectelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}subject")
    cartographicselement = etree.SubElement(subjectelement, "{http://www.loc.gov/mods/v3}cartographics")
    cartographicExtensionelement = etree.SubElement(cartographicselement, "{http://www.loc.gov/mods/v3}cartographicExtension")

    coordinateselement = etree.SubElement(cartographicExtensionelement, "{http://www.loc.gov/mods/v3}coordinates")
    coordinateselement.text = ' '.join(row.get("coordinates", '').split())

    scaleelement = etree.SubElement(cartographicExtensionelement, "{http://www.loc.gov/mods/v3}scale")
    scaleelement.text = ' '.join(row.get("scale", '').split())

    projectionelement = etree.SubElement(cartographicExtensionelement, "{http://www.loc.gov/mods/v3}projection")
    projectionelement.text = ' '.join(row.get("projection", '').split())

    #collection
    relatedItemelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}relatedItem", {"type":"host"})

    hosttitleInfoelement = etree.SubElement(relatedItemelement, "{http://www.loc.gov/mods/v3}titleInfo")
    hosttitleelement = etree.SubElement(hosttitleInfoelement, "{http://www.loc.gov/mods/v3}title")
    hosttitleelement.text = ' '.join(row.get("collection", '').split())

    hostoriginInfoelement = etree.SubElement(relatedItemelement, "{http://www.loc.gov/mods/v3}originInfo")
    hostdateCreatedelement = etree.SubElement(hostoriginInfoelement, "{http://www.loc.gov/mods/v3}dateCreated")
    hostdateCreatedelement.text = ' '.join(row.get("dateTextParent", '').split()).replace('.0','')

    hostidentifierelement = etree.SubElement(relatedItemelement, "{http://www.loc.gov/mods/v3}identifier", {"type":"local"})
    hostidentifierelement.text = ' '.join(row.get("callNumber", '').split())

    hostlocationelement = etree.SubElement(relatedItemelement, "{http://www.loc.gov/mods/v3}location")

    hostphysicalLocationelement = etree.SubElement(hostlocationelement, "{http://www.loc.gov/mods/v3}physicalLocation")
    hostphysicalLocationelement.text = ' '.join(row.get("repository", '').split())

    hosturlelement = etree.SubElement(hostlocationelement, "{http://www.loc.gov/mods/v3}url")
    hosturlelement.text = ' '.join(row.get("findingAid", '').split())

    hostholdingSimpleelement = etree.SubElement(hostlocationelement, "{http://www.loc.gov/mods/v3}holdingSimple")
    hostcopyInformationelement = etree.SubElement(hostholdingSimpleelement, "{http://www.loc.gov/mods/v3}copyInformation")
    hostshelfLocatorelement = etree.SubElement(hostcopyInformationelement, "{http://www.loc.gov/mods/v3}shelfLocator")

    shelfLocatorstring = ""

    if row.get("shelfLocator1", '') != "":
        shelfLocatorstring = ' '.join(row.get("shelfLocator1",'').split()) + ' ' + ' '.join(str(row.get("shelfLocator1ID",'')).split()).replace('.0','')
    if row.get("shelfLocator2", '') != "":
        shelfLocatorstring = shelfLocatorstring + ', ' + ' '.join(row.get("shelfLocator2",'').split()) + ' ' + ' '.join(str(row.get("shelfLocator2ID",'')).split()).replace('.0','')
    if row.get("shelfLocator3", '') != "":
        shelfLocatorstring = shelfLocatorstring + ', ' + ' '.join(row.get("shelfLocator3",'').split()) + ' ' + ' '.join(str(row.get("shelfLocator3ID",'')).split()).replace('.0','')

    hostshelfLocatorelement.text = ' '.join(shelfLocatorstring.split())

    #identifiers
    BDRPIDIdentifierelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}identifier", {"type":"local","displayLabel":"BDR_PID"})
    BDRPIDIdentifierelement.text = 'bdr:'+ ' '.join(row.get("identifierBDR", '').split()).lstrip('bdr')

    MODSIDIdentifierelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}identifier", {"type":"local","displayLabel":"MODS_ID"})
    MODSIDIdentifierelement.text = 'bdr'+ ' '.join(row.get("identifierBDR", '').split()).lstrip('bdr')

    #lastnote
    digitalObjectMadeelement = etree.SubElement(modstop, "{http://www.loc.gov/mods/v3}note", {"displayLabel":"Digital object made available by"})
    digitalObjectMadeelement.text = "Brown University Library, John Hay Library, University Archives and Manuscripts, Box A, Brown University, Providence, RI, 02912, U.S.A., (http://library.brown.edu/)"

    # start cleanup
    # remove any element tails
    for element in modstop.iter():
        element.tail = None

    # remove any line breaks or tabs in element text
        if element.text:
            if '\n' in element.text:
                element.text = element.text.replace('\n', '')
            if '\t' in element.text:
                element.text = element.text.replace('\t', '')

    # remove any remaining whitespace
    parser = etree.XMLParser(remove_blank_text=True, remove_comments=True, recover=True)
    treestring = etree.tostring(modstop)
    clean = etree.XML(treestring, parser)

    # remove recursively empty nodes
    # found here: https://stackoverflow.com/questions/12694091/python-lxml-how-to-remove-empty-repeated-tags
    def recursively_empty(e):
       if e.text:
           return False
       return all((recursively_empty(c) for c in e.iterchildren()))

    context = etree.iterwalk(clean)
    for action, elem in context:
        parent = elem.getparent()
        if recursively_empty(elem):
            parent.remove(elem)

    # remove nodes with blank attribute
    for element in clean.xpath(".//*[@*='']"):
        element.getparent().remove(element)

    # remove nodes with attribute "null"
    for element in clean.xpath(".//*[@*='null']"):
        element.getparent().remove(element)

    filename = row.get("identifierBDR", '')

    if filename == "":
        filename = "default" + str(rowindex)

    with open(output_path+'/'+filename + ".mods", 'wb') as f:
        f.write("<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n")
        f.write(etree.tostring(clean, pretty_print = True))
        print "Writing" + filename + ".mods"

    rowindex = rowindex + 1
    amountofrecords = amountofrecords + 1

if langissue == True:
    messageToUser("*Language Field Error*", "There were one or more issues with language fields in your spreadsheet. Please check your spelling in all language fields. You may also manually correct your XML file, consult the SupportedLanguages.xlsx file in the data folder for supported languages, and/or adjust the SupportedLanguages.xlsx spreadsheet to suit your project.")

errorfile.close()

if amountofrecords > 1:
    messageToUser("***Operation Complete***", str(amountofrecords) + " MODS records were written to folder " + output_path + ".")
else:
    messageToUser("***Operation Complete***", str(amountofrecords) + " MODS record was written to folder " + output_path + ".")
